
# -*- coding=utf-8 -*-
import requests
from bs4 import BeautifulSoup
#import os
from openpyxl import Workbook
from lxml import etree
import time

num0 = 1  # 用来计数，计算爬取的书一共有多少本
num1 = 1
num2 = 1 #设置三个计数的变量是为了将书籍、类别和作者对应起来

start_time = time.time()  # 计算爬虫爬取过程时间

# 第一页网页网址https://read.douban.com/columns/category/all?sort=hot&start=0
# 第二页网页网址https://read.douban.com/columns/category/all?sort=hot&start=10
# 第三页网页网址https://read.douban.com/columns/category/all?sort=hot&start=20
# ......发现规律了吗
url = 'https://read.douban.com/columns/category/all?sort=hot&start='

wb = Workbook()
ws = wb.active
ws.title = "豆瓣阅读全部专栏"
ws.cell(row=1, column=1).value = '全部专栏'
ws.cell(row=1, column=2).value = '作者'
ws.cell(row=1, column=3).value = '类别'

for i in range(0, 1760, 10):  # 这里的  range（初始，结束，间隔）
    # requests库用来向该网服务器发送请求，请求打开该网址链接。requests.get().content表示什么意思？
    html = requests.get('https://read.douban.com/columns/category/all?sort=hot&start=%d' % i).content
    # BeautifulSoup库解析获得的网页，第二个参数一定记住要写上‘lxml’，记住就行
    bsObj = BeautifulSoup(html, 'lxml')
    print('==============' + '第%d页' % (i / 10 + 1) + '==============')
    # 分析网页发现，每页有10本书，而<h4>标签正好只有10个。

    # 下面的for循环是爬取书籍名称
    h4_node_list = bsObj.find_all('h4')  # 这里返回的是h4标签的list列表。
    #对于为什么是find_all('h4'),以及下面的类似。都是查看页面源代码，然后找规律，参考BS4文档。
    for h4_node in h4_node_list:
    # 因为是列表，要用list[0]取出来<a>标签，在用<a>的string将文本取出来
        title = h4_node.contents[0].string
        title = '<<' + title + '>>'
        #title_all = title_all.append(title)
        print('第%d本书' % num0, title)
        num0 = num0 + 1
        ws.cell(row=num0, column=1).value = title

    # 下面的for循环是爬取对应书籍的类别
    category_node_list = bsObj.find_all("div","category")
    for category_node in category_node_list:
        category = category_node.contents[1].string
        print('第%d本书的类别' % num1, category)
        num1 = num1 + 1
        ws.cell(row=num1, column=3).value = category

    # 下面的for循环是爬取对应书籍的作者
    author_node_list = bsObj.find_all("div","author")
    #category_node_list = bsObj.find_all("div","category")
    for author_node in author_node_list:
        author = author_node.contents[1].string
        print('第%d本书的作者' % num2, author)
        num2 = num2 + 1
        ws.cell(row=num2, column=2).value = author

    wb.save('豆瓣阅读全部专栏' + '.xlsx')  #将书籍写入Excel表格

    time.sleep(2)
    #设置抓数据停顿时间为1秒，防止过于频繁访问该网站，被封

#下面用来统计爬虫用时及抓取书籍总数。
end_time = time.time()
duration_time = end_time - start_time
print('运行时间共：%.2f' %duration_time + '秒')
print('共抓到%d本书名' % (num1-1))
